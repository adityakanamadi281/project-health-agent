"""Loads Smartsheet-style project-plan exports (.xlsx) into a ProjectPlan.

Design goals (this is the "handle incomplete or messy data gracefully"
requirement in the assignment):

1. Header names are matched by *label*, not fixed column position, because the
   two sample workbooks already put the same field in different columns
   (e.g. "Area" is column 28 in one file and column 9 in the other).
2. Every raw cell is passed through tolerant parsers that turn Smartsheet junk
   values (``#UNPARSEABLE``, blanks, stray strings in numeric columns, "3d"
   duration strings, etc.) into ``None`` instead of raising or silently
   corrupting downstream math.
3. Every time a value could not be parsed where data was actually present, we
   record a DataQualityNote instead of failing — the RAG engine and LLM layer
   both surface these notes so a human knows the report was built on
   incomplete data, rather than pretending the number is trustworthy.
4. Missing sheets (e.g. no "Comments" sheet) degrade to an empty list rather
   than raising.
"""

from __future__ import annotations

import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from project_health_agent.models import (
    Comment,
    DataQualityNote,
    ProjectPlan,
    ProjectSummary,
    Task,
    TaskStatus,
)

UNPARSEABLE_MARKERS = {"#unparseable", "#n/a", "#ref!", "#value!", "n/a", "na", "-", ""}

# Canonical field -> acceptable header labels (case-insensitive, whitespace-normalized).
TASK_HEADER_ALIASES: dict[str, list[str]] = {
    "task_name": ["task name"],
    "level": ["level"],
    "phase_milestone": ["phase/milestone", "phase / milestone", "phase milestone"],
    "status": ["status"],
    "percent_complete": ["% complete", "percent complete", "%complete"],
    "start_date": ["start date"],
    "end_date": ["end date"],
    "baseline_start": ["baseline start"],
    "baseline_finish": ["baseline finish"],
    "variance": ["variance"],
    "duration": ["duration"],
    "priority": ["priority"],
    "critical": ["critical ?", "critical?", "critical"],
    "on_hold": ["on hold?", "on hold"],
    "not_applicable": ["not applicable?", "not applicable"],
    "owner": ["owner", "ownership"],
    "assigned_to": ["assigned to"],
    "area": ["area"],
    "status_comment": ["status comment"],
    "rag": ["rag"],
    "total_float": ["total float", "float"],
    "at_risk": ["at risk?", "at risk"],
    "predecessors": ["predecessors", "predecessor"],
}

SUMMARY_ROW_ALIASES: dict[str, list[str]] = {
    "project_name": ["project name"],
    "project_manager": ["project manager"],
    "project_start_date": ["project start date"],
    "project_end_date": ["project end date"],
    "not_started_count": ["not started"],
    "in_progress_count": ["in progress"],
    "completed_count": ["completed"],
    "on_hold_count": ["on hold"],
    "at_risk_label": ["at risk"],
    "project_stage": ["project stage"],
    "percent_complete": ["% complete", "percent complete"],
    "schedule_health_label": ["schedule health"],
    "as_of_date": ["today's date", "todays date", "as of date"],
    "duration_days": ["duration"],
    "project_status": ["project status"],
}


def _norm(label: object) -> str:
    return re.sub(r"\s+", " ", str(label or "")).strip().lower()


def _is_blankish(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and _norm(value) in UNPARSEABLE_MARKERS:
        return True
    return False


class _QualityTracker:
    """Accumulates (field, issue) counts so we can emit one note per issue type."""

    def __init__(self) -> None:
        self._counts: Counter[tuple[str, str]] = Counter()

    def flag(self, field: str, issue: str) -> None:
        self._counts[(field, issue)] += 1

    def to_notes(self) -> list[DataQualityNote]:
        return [
            DataQualityNote(field=field, issue=issue, count=count)
            for (field, issue), count in sorted(self._counts.items())
        ]


def _parse_float(value: object, field: str, qt: _QualityTracker) -> float | None:
    if _is_blankish(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        m = re.search(r"-?\d+(\.\d+)?", value)
        if m:
            return float(m.group())
    qt.flag(field, "non-numeric value ignored")
    return None


def _parse_duration_days(value: object, field: str, qt: _QualityTracker) -> float | None:
    """Smartsheet durations arrive as '262d', '0', or a plain number."""
    if _is_blankish(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        m = re.search(r"-?\d+(\.\d+)?", value)
        if m:
            return float(m.group())
    qt.flag(field, "unparseable duration ignored")
    return None


def _parse_variance_days(value: object, field: str, qt: _QualityTracker) -> int | None:
    if _is_blankish(value):
        return None
    if isinstance(value, (int, float)):
        return int(round(value))
    if isinstance(value, str):
        m = re.search(r"-?\d+", value)
        if m:
            return int(m.group())
    qt.flag(field, "unparseable variance ignored")
    return None


def _parse_date(value: object, field: str, qt: _QualityTracker) -> date | None:
    if _is_blankish(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    qt.flag(field, "unparseable date ignored")
    return None


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return _norm(value) in {"true", "yes", "y", "1"}
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def _parse_status(value: object) -> TaskStatus:
    if not isinstance(value, str):
        return TaskStatus.UNKNOWN
    v = _norm(value)
    mapping = {
        "not started": TaskStatus.NOT_STARTED,
        "in progress": TaskStatus.IN_PROGRESS,
        "completed": TaskStatus.COMPLETED,
        "complete": TaskStatus.COMPLETED,
        "on hold": TaskStatus.ON_HOLD,
    }
    return mapping.get(v, TaskStatus.UNKNOWN)


def _parse_percent(value: object, field: str, qt: _QualityTracker) -> float | None:
    if _is_blankish(value):
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        return v if v <= 1.0001 else v / 100.0
    if isinstance(value, str):
        m = re.search(r"-?\d+(\.\d+)?", value)
        if m:
            v = float(m.group())
            return v / 100.0 if v > 1.0001 else v
    qt.flag(field, "unparseable % complete ignored")
    return None


def _build_header_index(ws: Worksheet, header_row: int, aliases: dict[str, list[str]]) -> dict[str, int]:
    raw_headers = {
        col: _norm(ws.cell(row=header_row, column=col).value) for col in range(1, ws.max_column + 1)
    }
    index: dict[str, int] = {}
    for canonical, labels in aliases.items():
        for col, header in raw_headers.items():
            if header in labels:
                index[canonical] = col
                break
    return index


def _find_task_sheet(wb: openpyxl.Workbook) -> Worksheet:
    for name in wb.sheetnames:
        if name.lower() not in {"comments", "summary"}:
            return wb[name]
    raise ValueError("No task/plan sheet found in workbook (only Comments/Summary present).")


def _load_tasks(ws: Worksheet, qt: _QualityTracker) -> list[Task]:
    idx = _build_header_index(ws, header_row=1, aliases=TASK_HEADER_ALIASES)
    tasks: list[Task] = []
    current_phase: str | None = None

    for row in range(2, ws.max_row + 1):
        def cell(field: str) -> object:
            col = idx.get(field)
            return ws.cell(row=row, column=col).value if col else None

        name = cell("task_name")
        if _is_blankish(name):
            continue  # a genuinely empty row — not a data-quality issue, just skip

        phase_val = cell("phase_milestone")
        is_phase = not _is_blankish(phase_val)
        if is_phase:
            current_phase = str(phase_val).strip()

        level_raw = cell("level")
        level = int(level_raw) if isinstance(level_raw, (int, float)) else None

        task = Task(
            row_number=row,
            name=str(name).strip(),
            level=level,
            is_milestone_or_phase=is_phase,
            phase_name=current_phase,
            status=_parse_status(cell("status")),
            percent_complete=_parse_percent(cell("percent_complete"), "% Complete", qt),
            start_date=_parse_date(cell("start_date"), "Start Date", qt),
            end_date=_parse_date(cell("end_date"), "End Date", qt),
            baseline_start=_parse_date(cell("baseline_start"), "Baseline Start", qt),
            baseline_finish=_parse_date(cell("baseline_finish"), "Baseline Finish", qt),
            variance_days=_parse_variance_days(cell("variance"), "Variance", qt),
            duration_days=_parse_duration_days(cell("duration"), "Duration", qt),
            priority=(str(cell("priority")).strip() if not _is_blankish(cell("priority")) else None),
            critical=_parse_bool(cell("critical")),
            on_hold=_parse_bool(cell("on_hold")),
            not_applicable=_parse_bool(cell("not_applicable")),
            owner=(str(cell("owner")).strip() if not _is_blankish(cell("owner")) else None),
            assigned_to=(
                str(cell("assigned_to")).strip() if not _is_blankish(cell("assigned_to")) else None
            ),
            area=(str(cell("area")).strip() if not _is_blankish(cell("area")) else None),
            status_comment=(
                str(cell("status_comment")).strip()
                if not _is_blankish(cell("status_comment"))
                else None
            ),
            rag_raw=(str(cell("rag")).strip() if not _is_blankish(cell("rag")) else None),
            total_float=_parse_float(cell("total_float"), "Total Float", qt),
            at_risk=(str(cell("at_risk")).strip() if not _is_blankish(cell("at_risk")) else None),
            predecessors=(str(cell("predecessors")).strip() if not _is_blankish(cell("predecessors")) else None),
        )
        tasks.append(task)

    if not idx.get("task_name"):
        raise ValueError("Could not locate a 'Task Name' column — is this a supported export format?")

    return tasks


def _load_comments(wb: openpyxl.Workbook) -> list[Comment]:
    if "Comments" not in wb.sheetnames:
        return []
    ws = wb["Comments"]
    comments: list[Comment] = []
    for row in range(1, ws.max_row + 1):
        text = ws.cell(row=row, column=2).value
        if _is_blankish(text):
            continue
        comments.append(
            Comment(
                row_reference=(
                    str(ws.cell(row=row, column=1).value)
                    if not _is_blankish(ws.cell(row=row, column=1).value)
                    else None
                ),
                text=str(text).strip(),
                author=(
                    str(ws.cell(row=row, column=3).value)
                    if not _is_blankish(ws.cell(row=row, column=3).value)
                    else None
                ),
                timestamp=(
                    str(ws.cell(row=row, column=4).value)
                    if not _is_blankish(ws.cell(row=row, column=4).value)
                    else None
                ),
            )
        )
    return comments


def _load_summary(wb: openpyxl.Workbook, qt: _QualityTracker) -> ProjectSummary:
    if "Summary" not in wb.sheetnames:
        return ProjectSummary()
    ws = wb["Summary"]
    kv: dict[str, object] = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if label is None:
            continue
        kv[_norm(label)] = value

    def find(aliases: list[str]) -> object:
        for a in aliases:
            if a in kv:
                return kv[a]
        return None

    def as_int(v: object) -> int | None:
        try:
            return int(v) if v is not None and not _is_blankish(v) else None
        except (TypeError, ValueError):
            return None

    return ProjectSummary(
        project_name=(
            str(find(SUMMARY_ROW_ALIASES["project_name"]))
            if not _is_blankish(find(SUMMARY_ROW_ALIASES["project_name"]))
            else None
        ),
        project_manager=(
            str(find(SUMMARY_ROW_ALIASES["project_manager"]))
            if not _is_blankish(find(SUMMARY_ROW_ALIASES["project_manager"]))
            else None
        ),
        project_start_date=_parse_date(
            find(SUMMARY_ROW_ALIASES["project_start_date"]), "Summary: Project Start Date", qt
        ),
        project_end_date=_parse_date(
            find(SUMMARY_ROW_ALIASES["project_end_date"]), "Summary: Project End Date", qt
        ),
        not_started_count=as_int(find(SUMMARY_ROW_ALIASES["not_started_count"])),
        in_progress_count=as_int(find(SUMMARY_ROW_ALIASES["in_progress_count"])),
        completed_count=as_int(find(SUMMARY_ROW_ALIASES["completed_count"])),
        on_hold_count=as_int(find(SUMMARY_ROW_ALIASES["on_hold_count"])),
        at_risk_label=(
            str(find(SUMMARY_ROW_ALIASES["at_risk_label"]))
            if not _is_blankish(find(SUMMARY_ROW_ALIASES["at_risk_label"]))
            else None
        ),
        project_stage=(
            str(find(SUMMARY_ROW_ALIASES["project_stage"]))
            if not _is_blankish(find(SUMMARY_ROW_ALIASES["project_stage"]))
            else None
        ),
        percent_complete=_parse_percent(
            find(SUMMARY_ROW_ALIASES["percent_complete"]), "Summary: % Complete", qt
        ),
        schedule_health_label=(
            str(find(SUMMARY_ROW_ALIASES["schedule_health_label"]))
            if not _is_blankish(find(SUMMARY_ROW_ALIASES["schedule_health_label"]))
            else None
        ),
        as_of_date=_parse_date(find(SUMMARY_ROW_ALIASES["as_of_date"]), "Summary: As Of Date", qt),
        duration_days=_parse_duration_days(
            find(SUMMARY_ROW_ALIASES["duration_days"]), "Summary: Duration", qt
        ),
        project_status=(
            str(find(SUMMARY_ROW_ALIASES["project_status"]))
            if not _is_blankish(find(SUMMARY_ROW_ALIASES["project_status"]))
            else None
        ),
    )


def load_project_plan(path: str | Path) -> ProjectPlan:
    """Parse one .xlsx export into a normalized ProjectPlan.

    Raises FileNotFoundError / ValueError for genuinely unusable input (missing
    file, or no recognizable task-name column). Anything short of that degrades
    gracefully with DataQualityNotes rather than raising.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Project plan not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    qt = _QualityTracker()

    task_sheet = _find_task_sheet(wb)
    tasks = _load_tasks(task_sheet, qt)
    comments = _load_comments(wb)
    summary = _load_summary(wb, qt)

    if not summary.project_name:
        # Fall back to the top-level task name (row 2 in both sample files) or filename.
        top_level = next((t for t in tasks if t.level in (0, None)), None)
        summary.project_name = (top_level.name if top_level else None) or path.stem

    return ProjectPlan(
        source_file=path.name,
        sheet_name=task_sheet.title,
        summary=summary,
        tasks=tasks,
        comments=comments,
        data_quality_notes=qt.to_notes(),
    )
